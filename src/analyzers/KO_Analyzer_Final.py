import stanza
import numpy as np
import pandas as pd
from openpyxl import Workbook
import jiwer
from collections import defaultdict
from konlpy.tag import Okt
from typing import List, Dict, Tuple

class KoreanLinguisticAnalyzer:
    def __init__(self):
        """
        Initialize Korean language processing tools:
        - Stanza model for dependency parsing
        - Okt for morphological analysis
        - Custom particle handler for Korean-specific features
        """
        # Download the Korean Stanza model if not already available
        stanza.download('ko')
        self.nlp = stanza.Pipeline('ko', use_gpu=False)  # Force using CPU only
        
        self.okt = Okt()

        # Korean particle (조사) handling
        self.particles = set(['은', '는', '이', '가', '을', '를', '의', '에', '에서', '로', '으로'])

        # Korean sentence ending markers
        self.endings = set(['다', '요', '니다', '시오'])

    def preprocess_korean(self, text: str) -> str:
        """
        Preprocess Korean text:
        1. Normalize particles
        2. Handle sentence endings
        3. Adjust spacing for proper dependency parsing
        """
        # Morphological analysis
        morphs = self.okt.pos(text)

        # Handle particles
        processed_morphs = []
        for morph, pos in morphs:
            if pos == 'Josa':  # Particle
                if morph in self.particles:
                    processed_morphs.append(morph + '+PARTICLE')
            elif pos == 'Eomi':  # Ending
                if morph in self.endings:
                    processed_morphs.append(morph + '+END')
            else:
                processed_morphs.append(morph)

        return ' '.join(processed_morphs)

    def process_file(self, file_path: str):
        """Process input file with Korean-specific handling"""
        df = pd.read_excel(file_path)
        
        # Initialize output workbooks
        workbooks = {
            'poly': Workbook(),
            'wer': Workbook(),
            'hier': Workbook(),
            'linear': Workbook()
        }
        worksheets = {k: wb.active for k, wb in workbooks.items()}
        
        # Set worksheet titles
        for k, ws in worksheets.items():
            ws.title = f'Korean_{k.capitalize()}_Analysis'

        # Process each sentence pair
        for row_idx in range(len(df)):
            # Get base sentence and preprocess
            sentenceA = df.iloc[row_idx, 0]
            processed_A = self.preprocess_korean(sentenceA)
            docA = self.nlp(processed_A)

            # Write original sentence to all sheets
            for ws in worksheets.values():
                ws.cell(row=row_idx + 1, column=1, value=sentenceA)

            # Calculate base features
            plistA = self.dependency_tree_to_parent_list(docA)
            polyA = self.compute_polynomial(plistA)
            vectorA = self.polynomial_to_vector(polyA)
            hierA = self._get_korean_hierarchical_features(docA)
            linearA = self._get_korean_linear_features(docA)

            # Compare with other sentences
            for col_idx in range(1, len(df.columns)):
                sentenceB = df.iloc[row_idx, col_idx]
                if pd.isna(sentenceB):
                    continue

                processed_B = self.preprocess_korean(sentenceB)
                docB = self.nlp(processed_B)
                
                # Calculate metrics
                if not docB.sentences:
                    continue

                plistB = self.dependency_tree_to_parent_list(docB)
                polyB = self.compute_polynomial(plistB)
                vectorB = self.polynomial_to_vector(polyB)
                poly_dist = self.bipartite_distance(vectorA, vectorB)
                wer = self._calculate_korean_wer(sentenceA, sentenceB)
                hier_dist = self._calculate_hierarchical_distance(
                    hierA, 
                    self._get_korean_hierarchical_features(docB)
                )
                linear_dist = self._calculate_linear_distance(
                    linearA, 
                    self._get_korean_linear_features(docB)
                )

                # Store results
                worksheets['poly'].cell(row=row_idx + 1, column=col_idx + 1, value=poly_dist)
                worksheets['wer'].cell(row=row_idx + 1, column=col_idx + 1, value=wer)
                worksheets['hier'].cell(row=row_idx + 1, column=col_idx + 1, value=hier_dist)
                worksheets['linear'].cell(row=row_idx + 1, column=col_idx + 1, value=linear_dist)

        # Save results
        for k, wb in workbooks.items():
            wb.save(f'korean_output_{k}_analysis.xlsx')

    def dependency_tree_to_parent_list(self, doc):
        """
        Extract a parent-child relationship list from a Stanza document.
        """
        parent_list = []
        for sentence in doc.sentences:
            for word in sentence.words:
                parent_list.append((word.id, word.head, word.deprel))
        return parent_list

    def compute_polynomial(self, parent_list) -> List[Tuple[str, int]]:
        """
        Compute polynomial representation from parent-child relationships.
        """
        X = {f'x{i}': 0 for i in range(1, 38)}
        Y = {f'y{i}': 0 for i in range(1, 38)}

        label_to_poly = {
            'acl': 'x1', 'advcl': 'x2', 'advmod': 'x3', 'amod': 'x4',
            'appos': 'x5', 'aux': 'x6', 'case': 'x7', 'cc': 'x8',
            'ccomp': 'x9', 'clf': 'x10', 'compound': 'x11', 'conj': 'x12',
            'cop': 'x13', 'csubj': 'x14', 'dep': 'x15', 'det': 'x16',
            'discourse': 'x17', 'dislocated': 'x18', 'expl': 'x19', 'fixed': 'x20',
            'flat': 'x21', 'goeswith': 'x22', 'iobj': 'x23', 'list': 'x24',
            'mark': 'x25', 'nmod': 'x26', 'nsubj': 'x27', 'nummod': 'x28',
            'obj': 'x29', 'obl': 'x30', 'orphan': 'x31', 'parataxis': 'x32',
            'punct': 'x33', 'reparandum': 'x34', 'root': 'x35', 'vocative': 'x36',
            'xcomp': 'x37'
        }

        for child, parent, label in parent_list:
            if label in label_to_poly:
                term = label_to_poly[label]
                if child == parent:  # Root node
                    Y[term.replace('x', 'y')] += 1
                else:  # Non-root nodes
                    X[term] += 1

        poly = []
        for key, value in X.items():
            if value > 0:
                poly.append((key, value))
        for key, value in Y.items():
            if value > 0:
                poly.append((key, value))

        return poly

    def polynomial_to_vector(self, poly: List[Tuple[str, int]]) -> List[int]:
        """
        Convert polynomial representation to a vector.
        """
        vector = [0] * 75
        label_to_index = {f'x{i}': i for i in range(1, 38)}
        label_to_index.update({f'y{i}': i + 37 for i in range(1, 38)})

        for label, count in poly:
            index = label_to_index.get(label, -1)
            if index != -1:
                vector[index] = count
        return vector

    def bipartite_distance(self, vector1: List[int], vector2: List[int]) -> float:
        """
        Compute polynomial distance using Manhattan distance between the vectors.
        """
        return np.sum(np.abs(np.array(vector1) - np.array(vector2)))

    def _get_korean_hierarchical_features(self, doc) -> Dict:
        """
        Korean-specific hierarchical features:
        - Topic-comment structure
        - Embedded clause depth
        - Long-distance scrambling
        """
        def get_depth(token_id, sentence):
            children = [w for w in sentence.words if w.head == token_id]
            return max([get_depth(child.id, sentence) for child in children] or [0]) + 1

        # Check if there are any sentences and if the root exists
        if not doc.sentences or not any(word.head == 0 for word in doc.sentences[0].words):
            return {'depth': 0, 'branching': 0, 'long_dist': 0, 'topic_depth': 0}

        root = [word for word in doc.sentences[0].words if word.head == 0][0]
        
        features = {
            'depth': get_depth(root.id, doc.sentences[0]),
            'branching': np.mean([len([w for w in doc.sentences[0].words if w.head == word.id]) for word in doc.sentences[0].words]),
            'long_dist': sum(1 for word in doc.sentences[0].words if abs(word.id - word.head) > 3),
            'topic_depth': self._calculate_topic_depth(doc.sentences[0])
        }
        return features

    def _calculate_topic_depth(self, sentence) -> int:
        """Calculate depth of topic-comment structure"""
        topic_markers = ['은', '는']
        depth = 0
        for word in sentence.words:
            if any(marker in word.text for marker in topic_markers):
                current_depth = len([w for w in sentence.words if w.head == word.id])
                depth = max(depth, current_depth)
        return depth

    def _get_korean_linear_features(self, doc) -> Dict:
        """
        Korean-specific linear features:
        - SOV order adherence
        - Particle position patterns
        - Local word order variation
        """
        features = {
            'local_trans': self._get_korean_transitions(doc.sentences[0]),
            'particle_patterns': self._analyze_particle_patterns(doc.sentences[0]),
            'linear_dist': np.mean([abs(word.id - word.head) for word in doc.sentences[0].words if word.head != 0] or [0])
        }
        return features

    def _get_korean_transitions(self, sentence) -> float:
        """Calculate Korean-specific transition probabilities"""
        transitions = defaultdict(int)
        total = 0
        for word in sentence.words[:-1]:
            next_word = sentence.words[word.id] if word.id < len(sentence.words) - 1 else None
            if next_word:
                transitions[f"{word.upos}_{next_word.upos}"] += 1
                total += 1
        return len(transitions) / (total + 1e-10)

    def _analyze_particle_patterns(self, sentence) -> float:
        """Analyze patterns of particle usage"""
        particle_positions = []
        for word in sentence.words:
            if any(particle in word.text for particle in self.particles):
                particle_positions.append(word.id / len(sentence.words))
        return np.mean(particle_positions) if particle_positions else 0

    def _calculate_korean_wer(self, original: str, comparison: str) -> float:
        """
        Modified WER calculation focusing on retention rather than exact reproduction:
        - Only counts substitutions and insertions
        - Ignores deletions to focus on what is retained
        - Morpheme-based comparison for Korean
        - Particle-aware comparison
        """
        def levenshtein_retention(ref: List[str], hyp: List[str]) -> float:
            """
            Modified Levenshtein distance that focuses on retention:
            - No penalty for deletions
            - Counts substitutions and insertions
            """
            # Create matrix
            dp = [[0] * (len(hyp) + 1) for _ in range(len(ref) + 1)]
            
            # Initialize first row (for insertions)
            for j in range(len(hyp) + 1):
                dp[0][j] = j
            
            # Fill the matrix
            for i in range(1, len(ref) + 1):
                for j in range(1, len(hyp) + 1):
                    if ref[i-1] == hyp[j-1]:  # Match
                        dp[i][j] = dp[i-1][j-1]
                    else:
                        deletion = dp[i-1][j]      # Cost of 0 for deletion
                        substitution = dp[i-1][j-1] + 1
                        insertion = dp[i][j-1] + 1
                        dp[i][j] = min(deletion, substitution, insertion)
            
            # Calculate retention-focused score
            error_count = dp[len(ref)][len(hyp)]
            denominator = max(len(ref), len(hyp))  # Use larger length for normalization
            
            return error_count / denominator if denominator > 0 else 0.0

        # Get morphemes using Okt
        original_morphs = [morph for morph, _ in self.okt.pos(original)]
        comparison_morphs = [morph for morph, _ in self.okt.pos(comparison)]

        # Special handling for particles
        def normalize_particles(morphs):
            normalized = []
            for morph in morphs:
                if morph in self.particles:
                    normalized.append(morph + '+PARTICLE')
                else:
                    normalized.append(morph)
            return normalized

        # Normalize particles in both sequences
        original_morphs = normalize_particles(original_morphs)
        comparison_morphs = normalize_particles(comparison_morphs)
        
        # Calculate retention-focused WER
        return levenshtein_retention(original_morphs, comparison_morphs)

    def _calculate_hierarchical_distance(self, hierA: Dict, hierB: Dict) -> float:
        """Calculate distance between hierarchical features"""
        return sum(abs(hierA[key] - hierB[key]) for key in hierA)

    def _calculate_linear_distance(self, linearA: Dict, linearB: Dict) -> float:
        """Calculate distance between linear features"""
        return sum(abs(linearA[key] - linearB[key]) for key in linearA)

def main():
    analyzer = KoreanLinguisticAnalyzer()
    analyzer.process_file("input.xlsx")

if __name__ == "__main__":
    main()
